import os
import sys
import datetime

import jaydebeapi
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Border, Side
import main.query.oracle as oracle
import main.query.pgsql as pgsql
import main.query.tibero as tibero
import main.query.db2 as db2


# Path configuration


# Border styles configuration
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

BOLD_BORDER = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

# 병합된 셀 테두리 설정하는 함수
def apply_border_to_merged_cells(sheet, merged_range, border):
    min_col, min_row, max_col, max_row = range_boundaries(merged_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = border



def create_list_sheet(workbook, rdbms):

    # sheet 활성화 및 "테이블 목록"라는 이름으로 변경
    list_sheet = workbook.create_sheet("테이블 목록")
    workbook.active = workbook.index(list_sheet)
    
    
    
    # 컬럼 레이아웃 설정
    list_sheet.cell(row=1, column=1, value="NO")
    list_sheet.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    list_sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    list_sheet.cell(row=1, column=2, value="업무 영역")
    list_sheet.cell(row=1, column=2).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    list_sheet.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
    list_sheet.cell(row=1, column=3, value="TABLE ID(물리명)")
    list_sheet.cell(row=1, column=3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    list_sheet.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
    list_sheet.cell(row=1, column=4, value="TABLE 명(논리명)")
    list_sheet.cell(row=1, column=4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    list_sheet.cell(row=1, column=4).font = openpyxl.styles.Font(bold=True)
    list_sheet.cell(row=1, column=5, value="TABLE 적재 주기")
    list_sheet.cell(row=1, column=5).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    list_sheet.cell(row=1, column=5).font = openpyxl.styles.Font(bold=True)
    list_sheet.cell(row=1, column=6, value="작성자")
    list_sheet.cell(row=1, column=6).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    list_sheet.cell(row=1, column=6).font = openpyxl.styles.Font(bold=True)
    list_sheet.cell(row=1, column=7, value="작성일")
    list_sheet.cell(row=1, column=7).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    list_sheet.cell(row=1, column=7).font = openpyxl.styles.Font(bold=True)
    
    # 테이블 목록 가져오기
    if rdbms == "Oracle":
        tables = oracle.oracle_table_list()
    elif rdbms == "PostgreSQL":
        tables = pgsql.pgsql_table_list()
    elif rdbms == "MSSQL":
        raise NotImplementedError("MSSQL is not supported yet.")
    elif rdbms == "mariaDB":
        raise NotImplementedError("mariaDB is not supported yet.")
    elif rdbms == "Tibero":
        tables = tibero.tibero_table_list()
    elif rdbms == "DB2(LUW)":
        tables = db2.db2_table_list()
        
    
    # 테이블 목록을 엑셀 시트에 추가
    for i, table in enumerate(tables, start=2):
        list_sheet.cell(row=i, column=1, value=i-1)
        list_sheet.cell(row=i, column=2, value=" ")
        list_sheet.cell(row=i, column=3, value=f"{table[0]}")
        list_sheet.cell(row=i, column=4, value=f"{table[1]}")
        list_sheet.cell(row=i, column=5, value="예시 - 년/월/일/시간")
        list_sheet.cell(row=i, column=6, value="예시 - 홍길동")
        list_sheet.cell(row=i, column=7, value=datetime.datetime.now().strftime("%Y-%m-%d"))
    
    
    # 데이터 내 width 조절
    start_row = 2
    start_column = 1
    end_column = 7
    for col in range(start_column, end_column + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            
        
            # 각 열의 4행부터 아래로 스캔
            for row in workbook.active.iter_rows(min_row=start_row, min_col=col, max_col=col):
                cell = row[0]
                value = cell.value
                if isinstance(value, str):
                    max_length = max(max_length, len(value))
                elif value is not None:
                    max_length = max(max_length, len(str(value)))
        
            adjusted_width = (max_length + 2) * 1.2  # Adjust the width as needed 
            workbook.active.column_dimensions[col_letter].width = adjusted_width
    
    # 전체 레이아옷 테두리 설정
    min_row = workbook.active.min_row
    max_row = workbook.active.max_row
    min_col = workbook.active.min_column
    max_col = workbook.active.max_column
    
    for row in workbook.active.iter_rows(min_row=min_row
                                         , max_row=max_row
                                         , min_col=min_col
                                         , max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell.border = THIN_BORDER
                
    
    

def create_table_sheet(workbook, path, rdbms):

    # table 목록 가져오기
    if rdbms == "Oracle":
        tables = oracle.oracle_table_list()
    elif rdbms == "PostgreSQL":
        tables = pgsql.pgsql_table_list()
    elif rdbms == "MSSQL":
        raise NotImplementedError("MSSQL is not supported yet.")
    elif rdbms == "mariaDB":
        raise NotImplementedError("mariaDB is not supported yet.")
    elif rdbms == "Tibero":
        tables = tibero.tibero_table_list()
    elif rdbms == "DB2(LUW)":
        tables = db2.db2_table_list()
    
       
    # 각 테이블에 대한 시트 생성
    for tables in tables :
        
        tables_sheet = workbook.create_sheet(f"{tables[0]}") # table 물리명으로 sheet 생성성
        workbook.active = workbook.index(tables_sheet)
        
        # 컬럼 레이아웃 설정
        tables_sheet.merge_cells("A1:B1")
        tables_sheet["A1"] = "TARGET TABLE 물리명"
        tables_sheet["A1"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["A1"].font = openpyxl.styles.Font(bold=True) 
        
        tables_sheet.merge_cells("C1:H1")
        tables_sheet["C1"] = f"{tables[0]}" # table 물리명
        tables_sheet["C1"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["C1"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet.merge_cells("I1:J1")
        tables_sheet["I1"] = "TARGET TABLE 논리명"
        tables_sheet["I1"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["I1"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet.merge_cells("K1:P1")
        tables_sheet["K1"] = f"{tables[1]}" # table 논리명
        tables_sheet["K1"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["K1"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet.merge_cells("A2:H2")
        tables_sheet["A2"] = "TARGET"
        tables_sheet["A2"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["A2"].font = openpyxl.styles.Font(bold=True)    
        
        tables_sheet.merge_cells("I2:M2")
        tables_sheet["I2"] = "SOURCE"
        tables_sheet["I2"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["I2"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet.merge_cells("N2:N3")
        tables_sheet["N2"] = "변환로직"
        tables_sheet["N2"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["N2"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet.merge_cells("O2:O3")
        tables_sheet["O2"] = "전체/증분"
        tables_sheet["O2"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["O2"].font = openpyxl.styles.Font(bold=True)
        
        
        tables_sheet.merge_cells("P2:P3")
        tables_sheet["P2"] = "비고"
        tables_sheet["P2"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["P2"].font = openpyxl.styles.Font(bold=True)
        
        # set the layout column 
        # target layout
        tables_sheet["A3"] = "NO"
        tables_sheet["A3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["A3"].font = openpyxl.styles.Font(bold=True)  
        
        tables_sheet["B3"] = "TABLE_NAME"
        tables_sheet["B3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["B3"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet["C3"] = "TABLE_COMMENT"
        tables_sheet["C3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["C3"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet["D3"] = "COLUMN_NAME"
        tables_sheet["D3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["D3"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet["E3"] = "COLUMN_COMMENT"
        tables_sheet["E3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["E3"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet["F3"] = "DATA_TYPE"
        tables_sheet["F3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["F3"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet["G3"] = "PK"
        tables_sheet["G3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["G3"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet["H3"] = "NULL"
        tables_sheet["H3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["H3"].font = openpyxl.styles.Font(bold=True)
        
        # source layout
        tables_sheet["I3"] = "TABLE_NAME"
        tables_sheet["I3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["I3"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet["J3"] = "TABLE_COMMENT"
        tables_sheet["J3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["J3"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet["K3"] = "COLUMN_NAME"
        tables_sheet["K3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["K3"].font = openpyxl.styles.Font(bold=True)
            
        tables_sheet["L3"] = "COLUMN_COMMENT"
        tables_sheet["L3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["L3"].font = openpyxl.styles.Font(bold=True)
        
        tables_sheet["M3"] = "DATA_TYPE"
        tables_sheet["M3"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet["M3"].font = openpyxl.styles.Font(bold=True)
        
        if rdbms == "Oracle":
            table_columns = oracle.oracle_table_info(tables[0])
        elif rdbms == "PostgreSQL":
            table_columns = pgsql.pgsql_table_info(tables[0])
        elif rdbms == "MSSQL":
            raise NotImplementedError("MSSQL is not supported yet.")
        elif rdbms == "mariaDB":
            raise NotImplementedError("mariaDB is not supported yet.")
        elif rdbms == "Tibero":
            table_columns = tibero.tibero_table_info(tables[0])
        elif rdbms == "DB2(LUW)":
            table_columns = db2.db2_table_info(tables[0])
            
        
            # 테이블 목록을 엑셀 시트에 추가
        for i, table_column in enumerate(table_columns, start=4):
            tables_sheet.cell(row=i, column=1, value=i-3)
            tables_sheet.cell(row=i, column=2, value=f"{table_column[0]}")
            tables_sheet.cell(row=i, column=3, value=f"{table_column[1]}")
            tables_sheet.cell(row=i, column=4, value=f"{table_column[2]}")
            tables_sheet.cell(row=i, column=5, value=f"{table_column[3]}")
            tables_sheet.cell(row=i, column=6, value=f"{table_column[4]}")
            tables_sheet.cell(row=i, column=7, value=f"{table_column[5]}")
            tables_sheet.cell(row=i, column=8, value=f"{table_column[6]}")
        
        # footer cell layout
        row_end = len(table_columns) + 3
        
        # 초기 적재 규칙
        tables_sheet.merge_cells(f"A{row_end+1}:B{row_end+1}")
        tables_sheet[f"A{row_end+1}"] = "초기 적재 규칙"
        tables_sheet[f"A{row_end+1}"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet[f"A{row_end+1}"].font = openpyxl.styles.Font(bold=True) 
        
        # 입력칸 
        tables_sheet.merge_cells(f"C{row_end+1}:H{row_end+1}")
        tables_sheet[f"C{row_end+1}"] = " "
        tables_sheet[f"C{row_end+1}"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet[f"C{row_end+1}"].font = openpyxl.styles.Font(bold=True)
        
        # 전체/증분 적재 규칙
        tables_sheet.merge_cells(f"I{row_end+1}:J{row_end+1}")
        tables_sheet[f"I{row_end+1}"] = "전체/증분 적재 규칙"
        tables_sheet[f"I{row_end+1}"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet[f"I{row_end+1}"].font = openpyxl.styles.Font(bold=True)
        
        # 입력칸
        tables_sheet.merge_cells(f"K{row_end+1}:P{row_end+1}")
        tables_sheet[f"K{row_end+1}"] = " "
        tables_sheet[f"K{row_end+1}"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet[f"K{row_end+1}"].font = openpyxl.styles.Font(bold=True)
        
        # 비고
        tables_sheet.merge_cells(f"A{row_end+2}:B{row_end+2}")
        tables_sheet[f"A{row_end+2}"] = "비고"
        tables_sheet[f"A{row_end+2}"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet[f"A{row_end+2}"].font = openpyxl.styles.Font(bold=True)
        
        # 입력칸
        tables_sheet.merge_cells(f"C{row_end+2}:P{row_end+2}")
        tables_sheet[f"C{row_end+2}"] = " "
        tables_sheet[f"C{row_end+2}"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        tables_sheet[f"C{row_end+2}"].font = openpyxl.styles.Font(bold=True)
        
        # 전체 width 조절
        start_row = 4
        start_column = 2
        end_column = 16

        for col in range(start_column, end_column + 1):
            max_length = 15
            col_letter = get_column_letter(col)
            
        
            # 각 열의 4행부터 아래로 스캔
            for row in workbook.active.iter_rows(min_row=start_row, min_col=col, max_col=col):
                cell = row[0]
                value = cell.value
                if isinstance(value, str):
                    max_length = max(max_length, len(value))
                elif value is not None:
                    max_length = max(max_length, len(str(value)))
        
            adjusted_width = (max_length + 2) * 1.2  # Adjust the width as needed 
            workbook.active.column_dimensions[col_letter].width = adjusted_width
        
        
        
       # merged range setting
        merged_ranges = ["A1:B1"
                         , "C1:H1"
                         , "I1:J1"
                         , "K1:P1"
                         , "A2:H2"
                         , "I2:M2"
                         , "N2:N3"
                         , "O2:O3"
                         , "P2:P3"
                         , f"A{row_end+1}:B{row_end+1}"
                         , f"C{row_end+1}:H{row_end+1}"
                         , f"I{row_end+1}:J{row_end+1}"
                         , f"K{row_end+1}:P{row_end+1}"
                         , f"A{row_end+2}:B{row_end+2}"
                         , f"C{row_end+2}:P{row_end+2}"
                         ] 
        
        # Apply border to each merged range
        for merged_range in merged_ranges:
            apply_border_to_merged_cells(tables_sheet, merged_range, THIN_BORDER)
            
        # not merge cell 레이아옷 테두리 설정
        # 쿼리 결과 셀 부분 테두리 설정
        for row in workbook.active.iter_rows(min_row=3              
                                             , max_row=row_end      
                                             , min_col=1            
                                             , max_col=16):         
             for cell in row:
                cell.border = THIN_BORDER

    # Save Path configuration
    SAVE_PATH = os.path.join(path, "매핑 정의서.xlsx")
            
    # Save the workbook to the specified file
    workbook.save(SAVE_PATH)
    